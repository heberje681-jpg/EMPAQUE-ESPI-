"""
Módulo de CAPTURA para la conciliación de pallets ESPI.

A diferencia de core.py (que lee una LIQUIDACIÓN ya armada en Excel con
bloques por día en cada hoja), este módulo modela los datos directamente:

    Productor (nombre, lote, teléfono, empresa, huerta)
      └── Monitoreo (fecha, folio, kilos recibidos/empacados/merma/jugo)
            └── Pallet (# pallet, variedad, calibre, cajas, estado, mixto)
                  └── (opcional) asignado a un Embarque

    Embarque (chofer, teléfono, placas, destino, contacto de llegada, estatus)

Los pallets ya "saben" en qué monitoreo se capturaron (no hay que
buscarlos), así que solo falta cruzarlos contra el RCF (manifiestos reales)
para saber cuáles ya se cargaron/coinciden y cuáles faltan. Los embarques
permiten armar/deshacer viajes y saber cuántos pallets/cajas le salieron a
cada productor por viaje.

Todo vive en tablas (listas de dicts) que la app mantiene en memoria
(session_state o el store compartido de Google Sheets); este módulo son
funciones puras sobre esas tablas más lectura/escritura de un Excel de
respaldo (para "guardar" y "continuar después" sin depender de que Google
Sheets esté configurado).
"""
import re
from collections import defaultdict
from difflib import SequenceMatcher

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import core  # reutiliza extract_rcf_records / index_rcf / _norm / similitud de nombres

_norm = core._norm
_name_similarity = core._name_similarity
PRODUCTOR_MATCH_THRESHOLD = core.PRODUCTOR_MATCH_THRESHOLD

CALIBRES_ESTANDAR = ["6X8", "8", "9", "10", "12", "14", "16", "18", "20", "22", "24", "6X6"]

# Cajas estándar por pallet según calibre (ajustable — son los valores que
# describió Sherlyn; un pallet con línea extra puede llevar más cajas, por
# eso esto es solo una SUGERENCIA que se puede sobreescribir al capturar).
CAJAS_ESTANDAR_POR_CALIBRE = {
    "6X8": 40,
    "6X6": 75,
    "8": 225, "9": 225, "10": 225, "12": 225, "14": 225,
    "16": 225, "18": 225, "20": 225, "22": 225, "24": 225,
}
CAJAS_CON_LINEA_EXTRA = 240  # calibres 10-24 cuando llevan una línea de más

ESTADOS_PALLET = ["cuarto_frio", "cargando", "en_viaje", "entregado"]
ESTATUS_EMBARQUE = ["armando", "cargando", "en_viaje", "entregado"]


def sugerir_cajas(calibre, con_linea_extra=False):
    """Sugerencia de cajas por pallet según calibre; el usuario la puede
    cambiar al capturar, esto solo evita tecleo repetitivo."""
    cal = _norm(calibre)
    if cal in ("6X8",):
        return CAJAS_ESTANDAR_POR_CALIBRE["6X8"]
    if cal in ("6X6",):
        return CAJAS_ESTANDAR_POR_CALIBRE["6X6"]
    base = CAJAS_ESTANDAR_POR_CALIBRE.get(cal)
    if base is None:
        return None
    return CAJAS_CON_LINEA_EXTRA if con_linea_extra else base


# ---------------------------------------------------------------------------
# Estado inicial
# ---------------------------------------------------------------------------

def nuevo_estado():
    return dict(productores=[], monitoreos=[], pallets=[], embarques=[], _next_id=1)


def _next_id(estado):
    i = estado["_next_id"]
    estado["_next_id"] += 1
    return i


def agregar_productor(estado, nombre, lote, telefono="", empresa="", huerta=""):
    pid = _next_id(estado)
    estado["productores"].append(dict(
        id=pid, nombre=nombre.strip(), lote=int(lote),
        telefono=telefono or "", empresa=empresa or "", huerta=huerta or "",
    ))
    return pid


def agregar_monitoreo(estado, productor_id, fecha, folio, kilos_recibidos,
                       kilos_empacados, kilos_merma, kilos_jugo):
    mid = _next_id(estado)
    estado["monitoreos"].append(dict(
        id=mid, productor_id=productor_id, fecha=fecha, folio=folio,
        kilos_recibidos=kilos_recibidos, kilos_empacados=kilos_empacados,
        kilos_merma=kilos_merma, kilos_jugo=kilos_jugo,
    ))
    return mid


def agregar_pallets(estado, monitoreo_id, filas):
    """filas: lista de dicts con 'pallet', 'calibre', 'cajas' (opcionales:
    'variedad', 'mixto', 'notas_mixto'). Los pallets nuevos entran con
    estado 'cuarto_frio' (recién armados, aún no salen a ningún viaje)."""
    agregados = []
    for f in filas:
        if f.get("pallet") in (None, ""):
            continue
        pid = _next_id(estado)
        registro = dict(
            id=pid, monitoreo_id=monitoreo_id,
            pallet=int(f["pallet"]), variedad=f.get("variedad") or None,
            calibre=f.get("calibre"),
            cajas=f.get("cajas") if f.get("cajas") not in ("", None) else None,
            mixto=bool(f.get("mixto", False)), notas_mixto=f.get("notas_mixto") or "",
            estado="cuarto_frio", embarque_id=None,
        )
        estado["pallets"].append(registro)
        agregados.append(registro)
    return agregados


def actualizar_pallets_de_monitoreo(estado, monitoreo_id, filas):
    """Reemplaza TODOS los pallets de un monitoreo por lo que venga en
    `filas` (una fila por pallet). Así se arma/deshace/edita/borra desde una
    sola tabla editable en la UI, sin funciones separadas para cada caso.
    Los pallets que ya estaban asignados a un embarque conservan esa
    asignación si su # de pallet se mantiene; si un pallet se borra de la
    tabla, se borra también su asignación a cualquier embarque."""
    existentes = {p["pallet"]: p for p in pallets_de_monitoreo(estado, monitoreo_id)}
    ids_conservados = set()
    nuevos_o_actualizados = []

    for f in filas:
        if f.get("pallet") in (None, ""):
            continue
        num = int(f["pallet"])
        cajas = f.get("cajas") if f.get("cajas") not in ("", None) else None
        mixto = bool(f.get("mixto", False))
        notas_mixto = f.get("notas_mixto") or ""
        variedad = f.get("variedad") or None
        calibre = f.get("calibre")
        if num in existentes:
            reg = existentes[num]
            reg.update(calibre=calibre, cajas=cajas, variedad=variedad,
                       mixto=mixto, notas_mixto=notas_mixto)
            ids_conservados.add(reg["id"])
            nuevos_o_actualizados.append(reg)
        else:
            pid = _next_id(estado)
            reg = dict(id=pid, monitoreo_id=monitoreo_id, pallet=num, variedad=variedad,
                       calibre=calibre, cajas=cajas, mixto=mixto, notas_mixto=notas_mixto,
                       estado="cuarto_frio", embarque_id=None)
            estado["pallets"].append(reg)
            ids_conservados.add(pid)
            nuevos_o_actualizados.append(reg)

    # quitar de estado los pallets de este monitoreo que ya no vinieron en `filas`
    estado["pallets"] = [
        p for p in estado["pallets"]
        if p["monitoreo_id"] != monitoreo_id or p["id"] in ids_conservados
    ]
    return nuevos_o_actualizados


def productor_por_id(estado, pid):
    return next((p for p in estado["productores"] if p["id"] == pid), None)


def monitoreos_de(estado, productor_id):
    return [m for m in estado["monitoreos"] if m["productor_id"] == productor_id]


def pallets_de_monitoreo(estado, monitoreo_id):
    return [p for p in estado["pallets"] if p["monitoreo_id"] == monitoreo_id]


def pallets_de_productor(estado, productor_id):
    mids = {m["id"] for m in monitoreos_de(estado, productor_id)}
    return [p for p in estado["pallets"] if p["monitoreo_id"] in mids]


def productor_de_pallet(estado, pallet):
    m = next((m for m in estado["monitoreos"] if m["id"] == pallet["monitoreo_id"]), None)
    if m is None:
        return None
    return productor_por_id(estado, m["productor_id"])


# ---------------------------------------------------------------------------
# Embarques (viajes): armar/deshacer, chofer/placas/destino, resumen por productor
# ---------------------------------------------------------------------------

def crear_embarque(estado, fecha_salida, chofer="", telefono_chofer="", placas="",
                    destino="", contacto_llegada=""):
    eid = _next_id(estado)
    estado["embarques"].append(dict(
        id=eid, fecha_salida=fecha_salida, chofer=chofer, telefono_chofer=telefono_chofer,
        placas=placas, destino=destino, contacto_llegada=contacto_llegada, estatus="armando",
    ))
    return eid


def pallets_disponibles(estado):
    """Pallets en cuarto frío, no asignados todavía a ningún embarque."""
    return [p for p in estado["pallets"] if p["estado"] == "cuarto_frio" and p["embarque_id"] is None]


def pallets_de_embarque(estado, embarque_id):
    return [p for p in estado["pallets"] if p["embarque_id"] == embarque_id]


def asignar_pallets_a_embarque(estado, embarque_id, pallet_ids):
    asignados = []
    for p in estado["pallets"]:
        if p["id"] in pallet_ids:
            p["embarque_id"] = embarque_id
            p["estado"] = "cargando"
            asignados.append(p)
    return asignados


def quitar_pallet_de_embarque(estado, pallet_id):
    for p in estado["pallets"]:
        if p["id"] == pallet_id:
            p["embarque_id"] = None
            p["estado"] = "cuarto_frio"
            return p
    return None


def cambiar_estatus_embarque(estado, embarque_id, nuevo_estatus):
    for e in estado["embarques"]:
        if e["id"] == embarque_id:
            e["estatus"] = nuevo_estatus
            estado_pallet = {"armando": "cargando", "cargando": "cargando",
                              "en_viaje": "en_viaje", "entregado": "entregado"}[nuevo_estatus]
            for p in pallets_de_embarque(estado, embarque_id):
                p["estado"] = estado_pallet
            return e
    return None


def resumen_embarque_por_productor(estado, embarque_id):
    """Cuántos pallets y cajas de cada productor van en este embarque —
    exactamente lo que Sherlyn necesita entregarle a cada productor."""
    filas = defaultdict(lambda: dict(pallets=0, cajas=0))
    for p in pallets_de_embarque(estado, embarque_id):
        productor = productor_de_pallet(estado, p)
        nombre = productor["nombre"] if productor else "(sin productor)"
        filas[nombre]["pallets"] += 1
        filas[nombre]["cajas"] += p["cajas"] or 0
    return dict(filas)


# ---------------------------------------------------------------------------
# Conciliación contra RCF
# ---------------------------------------------------------------------------

def reconciliar_productor(estado, productor_id, rcf_index, rcf_records_by_lote):
    productor = productor_por_id(estado, productor_id)
    lote = productor["lote"]
    nombre = productor["nombre"]
    monitoreos_by_id = {m["id"]: m for m in monitoreos_de(estado, productor_id)}
    pallets = pallets_de_productor(estado, productor_id)

    matched, mismatched, not_found = [], [], []
    listed_pallet_numbers = set()

    for p in pallets:
        listed_pallet_numbers.add(p["pallet"])
        candidates = rcf_index.get((lote, p["pallet"]), [])
        candidates = sorted(candidates, key=lambda rc: -_name_similarity(rc["productor"], nombre))
        if not candidates:
            not_found.append(p)
            continue
        rec = candidates[0]
        p = dict(p)  # copia para no mutar el estado original
        p["rcf_match"] = rec
        p["monitoreo"] = monitoreos_by_id.get(p["monitoreo_id"])
        same_calibre = (p["calibre"] is None) or (_norm(rec["calibre"]) == _norm(p["calibre"]))
        same_cajas = (p["cajas"] is None) or (rec["cajas"] == p["cajas"])
        if same_calibre and same_cajas:
            matched.append(p)
        else:
            mismatched.append(p)

    for p in not_found:
        p["monitoreo"] = monitoreos_by_id.get(p["monitoreo_id"])

    surplus = []
    for rec in rcf_records_by_lote.get(lote, []):
        if rec["pallet"] in listed_pallet_numbers:
            continue
        if nombre and _name_similarity(rec["productor"], nombre) < PRODUCTOR_MATCH_THRESHOLD:
            continue
        surplus.append(rec)

    by_calibre = defaultdict(lambda: dict(listados=0, cajas_listadas=0, encontrados=0,
                                           cajas_confirmadas=0, con_diferencia=0,
                                           no_encontrados=0, faltantes_pallets=[]))
    for p in pallets:
        cal = _norm(p["calibre"]) or "SIN CALIBRE"
        by_calibre[cal]["listados"] += 1
        by_calibre[cal]["cajas_listadas"] += p["cajas"] or 0
    for p in matched:
        cal = _norm(p["calibre"]) or "SIN CALIBRE"
        by_calibre[cal]["encontrados"] += 1
        by_calibre[cal]["cajas_confirmadas"] += p["rcf_match"]["cajas"] or 0
    for p in mismatched:
        cal = _norm(p["calibre"]) or "SIN CALIBRE"
        by_calibre[cal]["con_diferencia"] += 1
        by_calibre[cal]["cajas_confirmadas"] += p["rcf_match"]["cajas"] or 0
    for p in not_found:
        cal = _norm(p["calibre"]) or "SIN CALIBRE"
        by_calibre[cal]["no_encontrados"] += 1
        by_calibre[cal]["faltantes_pallets"].append(p["pallet"])

    return dict(
        productor=nombre, lote=lote, total_listados=len(pallets),
        matched=matched, mismatched=mismatched, not_found=not_found, surplus=surplus,
        by_calibre=dict(by_calibre),
    )


def reconciliar_todos(estado, rcf_index, rcf_records_by_lote):
    return {p["id"]: reconciliar_productor(estado, p["id"], rcf_index, rcf_records_by_lote)
            for p in estado["productores"]}


def cargar_rcf(rcf_path):
    wb_rcf = openpyxl.load_workbook(rcf_path, data_only=True)
    records, skipped = core.extract_rcf_records(wb_rcf)
    idx = core.index_rcf(records)
    by_lote = defaultdict(list)
    for r in records:
        by_lote[r["lote"]].append(r)
    return idx, by_lote, skipped


# ---------------------------------------------------------------------------
# Guardar / continuar (Excel de respaldo del estado capturado)
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_table(ws, start_row, headers, rows):
    for i, h in enumerate(headers):
        c = ws.cell(row=start_row, column=1 + i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = BORDER
    for r, row in enumerate(rows, start=start_row + 1):
        for i, v in enumerate(row):
            ws.cell(row=r, column=1 + i, value=v).border = BORDER
    for i in range(len(headers)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(1 + i)].width = 18
    return start_row + 1 + len(rows)


def exportar_captura(estado, path):
    """Excel de respaldo: tablas planas de productores/monitoreos/pallets,
    pensado para volver a subirlo después y seguir capturando donde quedó."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Productores")
    _write_table(ws, 1, ["id", "nombre", "lote"],
                 [[p["id"], p["nombre"], p["lote"]] for p in estado["productores"]])

    ws = wb.create_sheet("Monitoreos")
    _write_table(ws, 1, ["id", "productor_id", "fecha", "folio", "kilos_recibidos",
                          "kilos_empacados", "kilos_merma", "kilos_jugo"],
                 [[m["id"], m["productor_id"], m["fecha"], m["folio"], m["kilos_recibidos"],
                   m["kilos_empacados"], m["kilos_merma"], m["kilos_jugo"]]
                  for m in estado["monitoreos"]])

    ws = wb.create_sheet("Pallets")
    _write_table(ws, 1, ["id", "monitoreo_id", "pallet", "calibre", "cajas"],
                 [[p["id"], p["monitoreo_id"], p["pallet"], p["calibre"], p["cajas"]]
                  for p in estado["pallets"]])

    wb.save(path)


def importar_captura(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    estado = nuevo_estado()
    max_id = 0

    def rows(sheet_name):
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2):
            vals = [c.value for c in row]
            if all(v is None for v in vals):
                continue
            yield dict(zip(headers, vals))

    for r in rows("Productores"):
        estado["productores"].append(dict(id=int(r["id"]), nombre=r["nombre"], lote=int(r["lote"])))
        max_id = max(max_id, int(r["id"]))
    for r in rows("Monitoreos"):
        estado["monitoreos"].append(dict(
            id=int(r["id"]), productor_id=int(r["productor_id"]), fecha=r["fecha"], folio=r["folio"],
            kilos_recibidos=r["kilos_recibidos"], kilos_empacados=r["kilos_empacados"],
            kilos_merma=r["kilos_merma"], kilos_jugo=r["kilos_jugo"]))
        max_id = max(max_id, int(r["id"]))
    for r in rows("Pallets"):
        estado["pallets"].append(dict(
            id=int(r["id"]), monitoreo_id=int(r["monitoreo_id"]), pallet=int(r["pallet"]),
            calibre=r["calibre"], cajas=r["cajas"]))
        max_id = max(max_id, int(r["id"]))

    estado["_next_id"] = max_id + 1
    return estado


# ---------------------------------------------------------------------------
# Reporte final de conciliación (Excel legible, por productor)
# ---------------------------------------------------------------------------

FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")


def exportar_reporte_conciliacion(estado, resultados, path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_resumen = wb.create_sheet("Resumen")
    _write_table(ws_resumen, 1,
                 ["Productor", "Lote", "Pallets listados", "Encontrados", "Con diferencia",
                  "No encontrados", "Sobrantes en RCF"],
                 [[r["productor"], r["lote"], r["total_listados"], len(r["matched"]),
                   len(r["mismatched"]), len(r["not_found"]), len(r["surplus"])]
                  for r in resultados.values()])

    for pid, r in resultados.items():
        nombre_hoja = re.sub(r"[\\/*?:\[\]]", "", str(r["productor"]))[:28] or f"Lote {r['lote']}"
        ws = wb.create_sheet(nombre_hoja)
        ws.cell(row=1, column=1, value=f"LOTE {r['lote']} — {r['productor']}").font = Font(bold=True, size=13)

        row = 3
        headers = ["CALIBRE", "PALLETS LISTADOS", "CAJAS LISTADAS", "PALLETS ENCONTRADOS",
                   "CAJAS CONFIRMADAS (RCF)", "CON DIFERENCIA", "NO ENCONTRADOS", "# PALLETS FALTANTES"]
        cal_rows = []
        for cal in sorted(r["by_calibre"].keys()):
            d = r["by_calibre"][cal]
            cal_rows.append([cal, d["listados"], d["cajas_listadas"], d["encontrados"],
                              d["cajas_confirmadas"], d["con_diferencia"], d["no_encontrados"],
                              ", ".join(str(x) for x in d["faltantes_pallets"]) or "—"])
        row = _write_table(ws, row, headers, cal_rows) + 2

        ws.cell(row=row, column=1, value="Detalle de pallets capturados").font = Font(bold=True)
        row += 1
        det_headers = ["# Pallet", "Calibre", "Cajas", "Monitoreo (fecha)", "Estado", "Manifiesto RCF"]
        det_rows = []
        for p in r["matched"]:
            det_rows.append([p["pallet"], p["calibre"], p["cajas"],
                              p["monitoreo"]["fecha"] if p.get("monitoreo") else "", "OK",
                              p["rcf_match"]["manifiesto"]])
        for p in r["mismatched"]:
            det_rows.append([p["pallet"], p["calibre"], p["cajas"],
                              p["monitoreo"]["fecha"] if p.get("monitoreo") else "", "CON DIFERENCIA",
                              p["rcf_match"]["manifiesto"]])
        for p in r["not_found"]:
            det_rows.append([p["pallet"], p["calibre"], p["cajas"],
                              p["monitoreo"]["fecha"] if p.get("monitoreo") else "", "NO ENCONTRADO", ""])
        start = row
        row = _write_table(ws, row, det_headers, det_rows)
        for i, p in enumerate(r["matched"]):
            for col in range(1, 7):
                ws.cell(row=start + 1 + i, column=col).fill = FILL_OK
        offset = len(r["matched"])
        for i, p in enumerate(r["mismatched"]):
            for col in range(1, 7):
                ws.cell(row=start + 1 + offset + i, column=col).fill = FILL_WARN
        offset += len(r["mismatched"])
        for i, p in enumerate(r["not_found"]):
            for col in range(1, 7):
                ws.cell(row=start + 1 + offset + i, column=col).fill = FILL_BAD
        row += 2

        if r["surplus"]:
            ws.cell(row=row, column=1,
                    value="Pallets en RCF de este lote/productor que NO se capturaron en ningún monitoreo:"
                    ).font = Font(bold=True)
            row += 1
            row = _write_table(ws, row, ["# Pallet", "Calibre", "Cajas", "Productor (RCF)", "Manifiesto"],
                                [[x["pallet"], x["calibre"], x["cajas"], x["productor"], x["manifiesto"]]
                                 for x in sorted(r["surplus"], key=lambda x: x["pallet"])])

    wb.save(path)
