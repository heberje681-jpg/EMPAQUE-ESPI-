"""
Capa de acceso a datos — base de datos compartida (SQLite).

Todas las funciones abren su propia conexión corta (sqlite3 + WAL) para que
varias personas puedan leer/escribir desde distintas sesiones de Streamlit
sin pisarse. Si en el futuro se migra a Postgres/Supabase, solo hay que
reescribir esta capa (execute/query) — el resto de la app no debería
cambiar, porque todo pasa por estas funciones.
"""
import sqlite3
from contextlib import contextmanager
from datetime import datetime, date
from pathlib import Path

DB_PATH = Path(__file__).parent / "espi.db"

# Sugerencia de cajas por calibre (editable por el usuario al capturar).
CALIBRE_CAJAS_SUGERIDAS = {
    "8": 225, "9": 225, "10": 225, "12": 225, "14": 225, "16": 225,
    "18": 225, "20": 225, "22": 225, "24": 225,
    "6X6": 75, "6X8": 40,
}
CALIBRES = list(CALIBRE_CAJAS_SUGERIDAS.keys())

ESTADOS_PALLET = ["EN CUARTO FRÍO", "CARGANDO", "EN VIAJE", "ENTREGADO"]
ESTADOS_EMBARQUE = ["ARMANDO", "CARGANDO", "EN VIAJE", "ENTREGADO"]


@contextmanager
def _conn():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db():
    with _conn() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS productores (
            id INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            lote INTEGER,
            telefono TEXT,
            empresa TEXT,
            huerta TEXT
        )""")
        c.execute("""
        CREATE TABLE IF NOT EXISTS pallets (
            id INTEGER PRIMARY KEY,
            numero INTEGER NOT NULL,
            productor_id INTEGER REFERENCES productores(id),
            variedad TEXT,
            calibre TEXT,
            cajas INTEGER,
            organico INTEGER DEFAULT 0,
            mixto_notas TEXT,
            estado TEXT DEFAULT 'EN CUARTO FRÍO',
            embarque_id INTEGER REFERENCES embarques(id),
            fecha_creacion TEXT,
            activo INTEGER DEFAULT 1
        )""")
        c.execute("""
        CREATE TABLE IF NOT EXISTS embarques (
            id INTEGER PRIMARY KEY,
            fecha TEXT,
            chofer TEXT,
            telefono_chofer TEXT,
            placas TEXT,
            destino TEXT,
            contacto_llegada TEXT,
            estado TEXT DEFAULT 'ARMANDO',
            notas TEXT
        )""")
        c.execute("""
        CREATE TABLE IF NOT EXISTS remisiones (
            id INTEGER PRIMARY KEY,
            productor_id INTEGER REFERENCES productores(id),
            fecha TEXT,
            huerta TEXT,
            num_camiones INTEGER,
            cajas INTEGER,
            kilos REAL,
            notas TEXT
        )""")


# ---------------------------------------------------------------------------
# Productores
# ---------------------------------------------------------------------------

def crear_productor(nombre, lote=None, telefono="", empresa="", huerta=""):
    with _conn() as c:
        cur = c.execute(
            "INSERT INTO productores (nombre, lote, telefono, empresa, huerta) VALUES (?,?,?,?,?)",
            (nombre.strip(), lote, telefono, empresa, huerta))
        return cur.lastrowid


def listar_productores():
    with _conn() as c:
        return [dict(r) for r in c.execute("SELECT * FROM productores ORDER BY nombre")]


def productor(pid):
    with _conn() as c:
        r = c.execute("SELECT * FROM productores WHERE id=?", (pid,)).fetchone()
        return dict(r) if r else None


# ---------------------------------------------------------------------------
# Pallets
# ---------------------------------------------------------------------------

def crear_pallet(numero, productor_id, variedad, calibre, cajas, organico=False, mixto_notas=""):
    with _conn() as c:
        cur = c.execute(
            """INSERT INTO pallets (numero, productor_id, variedad, calibre, cajas, organico,
                                     mixto_notas, estado, fecha_creacion, activo)
               VALUES (?,?,?,?,?,?,?,?,?,1)""",
            (numero, productor_id, variedad, calibre, cajas, int(organico), mixto_notas,
             ESTADOS_PALLET[0], datetime.now().isoformat(timespec="seconds")))
        return cur.lastrowid


def listar_pallets(solo_activos=True, embarque_id=None, productor_id=None, estado=None):
    q = "SELECT p.*, pr.nombre AS productor_nombre, pr.lote AS lote FROM pallets p " \
        "LEFT JOIN productores pr ON pr.id = p.productor_id WHERE 1=1"
    params = []
    if solo_activos:
        q += " AND p.activo=1"
    if embarque_id is not None:
        q += " AND p.embarque_id=?"
        params.append(embarque_id)
    if productor_id is not None:
        q += " AND p.productor_id=?"
        params.append(productor_id)
    if estado is not None:
        q += " AND p.estado=?"
        params.append(estado)
    q += " ORDER BY p.numero"
    with _conn() as c:
        return [dict(r) for r in c.execute(q, params)]


def pallet(pid):
    with _conn() as c:
        r = c.execute("SELECT * FROM pallets WHERE id=?", (pid,)).fetchone()
        return dict(r) if r else None


def actualizar_pallet(pid, **campos):
    if not campos:
        return
    sets = ", ".join(f"{k}=?" for k in campos)
    with _conn() as c:
        c.execute(f"UPDATE pallets SET {sets} WHERE id=?", (*campos.values(), pid))


def deshacer_pallet(pid, nuevas_partes):
    """Divide un pallet en 2+ partes nuevas y desactiva el original.
    nuevas_partes: lista de dicts con numero, productor_id, variedad, calibre, cajas."""
    orig = pallet(pid)
    if not orig:
        return []
    nuevos_ids = []
    for parte in nuevas_partes:
        nid = crear_pallet(
            parte.get("numero", orig["numero"]),
            parte.get("productor_id", orig["productor_id"]),
            parte.get("variedad", orig["variedad"]),
            parte.get("calibre", orig["calibre"]),
            parte["cajas"],
            organico=orig["organico"],
            mixto_notas=f"Dividido del pallet #{orig['numero']} (id {orig['id']})",
        )
        nuevos_ids.append(nid)
    actualizar_pallet(pid, activo=0, mixto_notas=(orig["mixto_notas"] or "") +
                       f" | Dividido en pallets {nuevos_ids}")
    return nuevos_ids


def armar_pallet(pids, numero_resultante, productor_id, variedad, calibre):
    """Fusiona 2+ pallets en uno nuevo (suma cajas) y desactiva los originales."""
    partes = [pallet(pid) for pid in pids if pallet(pid)]
    if not partes:
        return None
    total_cajas = sum(p["cajas"] or 0 for p in partes)
    origenes = ", ".join(f"#{p['numero']}" for p in partes)
    nid = crear_pallet(numero_resultante, productor_id, variedad, calibre, total_cajas,
                        mixto_notas=f"Armado a partir de pallets {origenes}")
    for p in partes:
        actualizar_pallet(p["id"], activo=0,
                           mixto_notas=(p["mixto_notas"] or "") + f" | Fusionado en pallet #{numero_resultante}")
    return nid


# ---------------------------------------------------------------------------
# Embarques
# ---------------------------------------------------------------------------

def crear_embarque(fecha, chofer="", telefono_chofer="", placas="", destino="",
                    contacto_llegada="", notas=""):
    with _conn() as c:
        cur = c.execute(
            """INSERT INTO embarques (fecha, chofer, telefono_chofer, placas, destino,
                                       contacto_llegada, estado, notas)
               VALUES (?,?,?,?,?,?,?,?)""",
            (fecha, chofer, telefono_chofer, placas, destino, contacto_llegada,
             ESTADOS_EMBARQUE[0], notas))
        return cur.lastrowid


def listar_embarques():
    with _conn() as c:
        return [dict(r) for r in c.execute("SELECT * FROM embarques ORDER BY id DESC")]


def embarque(eid):
    with _conn() as c:
        r = c.execute("SELECT * FROM embarques WHERE id=?", (eid,)).fetchone()
        return dict(r) if r else None


def actualizar_embarque(eid, **campos):
    if not campos:
        return
    sets = ", ".join(f"{k}=?" for k in campos)
    with _conn() as c:
        c.execute(f"UPDATE embarques SET {sets} WHERE id=?", (*campos.values(), eid))


def asignar_pallet_a_embarque(pallet_id, embarque_id):
    actualizar_pallet(pallet_id, embarque_id=embarque_id, estado="CARGANDO")


def quitar_pallet_de_embarque(pallet_id):
    actualizar_pallet(pallet_id, embarque_id=None, estado="EN CUARTO FRÍO")


def resumen_embarque_por_productor(embarque_id):
    pallets = listar_pallets(embarque_id=embarque_id)
    resumen = {}
    for p in pallets:
        key = p["productor_nombre"] or "SIN PRODUCTOR"
        d = resumen.setdefault(key, dict(pallets=0, cajas=0))
        d["pallets"] += 1
        d["cajas"] += p["cajas"] or 0
    return resumen


# ---------------------------------------------------------------------------
# Remisiones (registro de recepción de camión) — base para el siguiente módulo
# ---------------------------------------------------------------------------

def crear_remision(productor_id, fecha, huerta, num_camiones, cajas, kilos, notas=""):
    with _conn() as c:
        cur = c.execute(
            """INSERT INTO remisiones (productor_id, fecha, huerta, num_camiones, cajas, kilos, notas)
               VALUES (?,?,?,?,?,?,?)""",
            (productor_id, fecha, huerta, num_camiones, cajas, kilos, notas))
        return cur.lastrowid


def listar_remisiones(productor_id=None):
    q = "SELECT r.*, pr.nombre AS productor_nombre FROM remisiones r " \
        "LEFT JOIN productores pr ON pr.id = r.productor_id WHERE 1=1"
    params = []
    if productor_id is not None:
        q += " AND r.productor_id=?"
        params.append(productor_id)
    q += " ORDER BY r.fecha DESC, r.id DESC"
    with _conn() as c:
        return [dict(r) for r in c.execute(q, params)]


init_db()


# ---------------------------------------------------------------------------
# Respaldo / restauración manual (Excel) — mientras no haya una base en la
# nube, esto es lo que protege los datos de un reinicio de la app.
# ---------------------------------------------------------------------------

_TABLAS_RESPALDO = ["productores", "embarques", "pallets", "remisiones"]


def exportar_backup(path):
    """Guarda todas las tablas en un Excel (una hoja por tabla), tal cual,
    para poder restaurarlas después con importar_backup()."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    with _conn() as c:
        for tabla in _TABLAS_RESPALDO:
            filas = [dict(r) for r in c.execute(f"SELECT * FROM {tabla}")]
            ws = wb.create_sheet(tabla)
            if filas:
                columnas = list(filas[0].keys())
                ws.append(columnas)
                for f in filas:
                    ws.append([f[col] for col in columnas])
            else:
                # conservar el encabezado aunque la tabla esté vacía
                cols = [d[1] for d in c.execute(f"PRAGMA table_info({tabla})")]
                ws.append(cols)
    wb.save(path)


def importar_backup(path):
    """Reemplaza el contenido actual de las tablas con lo que traiga el
    Excel de respaldo (mismo formato que genera exportar_backup)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    with _conn() as c:
        c.execute("PRAGMA foreign_keys=OFF")
        # borrar en orden hijo -> padre
        for tabla in ["pallets", "remisiones", "embarques", "productores"]:
            c.execute(f"DELETE FROM {tabla}")
        # insertar en orden padre -> hijo
        for tabla in _TABLAS_RESPALDO:
            if tabla not in wb.sheetnames:
                continue
            ws = wb[tabla]
            filas = list(ws.iter_rows(values_only=True))
            if not filas:
                continue
            headers = filas[0]
            placeholders = ", ".join("?" for _ in headers)
            cols = ", ".join(headers)
            for row in filas[1:]:
                if all(v is None for v in row):
                    continue
                c.execute(f"INSERT INTO {tabla} ({cols}) VALUES ({placeholders})", row)
        c.execute("PRAGMA foreign_keys=ON")
